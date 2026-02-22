"""
管理后台蓝图（Admin Backend Blueprint）。

功能：
- 管理员登录（AES 加密密码传输）
- 仪表盘统计（销售额、订单数、用户数、商品数）
- 商品增删改查 + 上下架切换
- 订单管理（状态更新）
- 用户列表
- 储值券生成
- 批量商品导入（Excel / JSON URL）
- 系统设置

相关漏洞：
- V-Admin-AES：前端硬编码 AES 密钥加密管理员密码，可被逆向破解
- V-SSRF：批量导入功能直接请求用户提供的 URL，未做校验

依赖：openpyxl（可选）用于解析 Excel 批量导入文件
"""

import io
import logging
import os
import urllib.request
from datetime import datetime, timedelta

# 后台管理蓝图：商品、订单、代金券与批量导入能力。

from flask import Blueprint, Response, flash, jsonify, redirect, render_template, request, session, url_for
from sqlalchemy import func
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import joinedload, selectinload

from app.models.db import Admin, Goods, Order, User, Voucher, db, GOODS_ON_SALE, GOODS_OFF_SALE, VOUCHER_UNUSED
from app.utils.tools import admin_auth, generate_uuid_hex, get_order_status_meta, is_admin_login, unique_filename

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None


admin_bp = Blueprint("admin", __name__, url_prefix="/admin")
logger = logging.getLogger(__name__)


def _upload_dir() -> str:
    # 管理后台上传目录统一入口，确保目录存在。
    path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")
    os.makedirs(path, exist_ok=True)
    return path


def _save_uploaded_image(file_storage):
    # 图片上传仅负责存储与回写路径，校验在业务层处理。
    if not file_storage or not file_storage.filename:
        return ""
    filename = unique_filename(file_storage.filename)
    file_storage.save(os.path.join(_upload_dir(), filename))
    return "/uploads/" + filename


def _commit_or_flash(success_msg: str, error_log: str, redirect_endpoint: str):
    # 后台表单提交统一事务与提示处理，减少重复代码。
    try:
        db.session.commit()
        flash(success_msg, "success")
    except SQLAlchemyError:
        db.session.rollback()
        logger.exception(error_log)
        flash("操作失败，请稍后重试", "danger")
    return redirect(url_for(redirect_endpoint))


@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    """
    管理员登录页面（F-402）
    漏洞：V-Admin-AES（保留）
    """
    if request.method == "POST":
        username = request.form.get("username")
        encrypted_password = request.form.get("password")

        if not username or not encrypted_password:
            flash("请输入用户名和密码", "danger")
            return render_template("admin/login.html")
        if admin_auth(username, encrypted_password):
            session["admin_logged_in"] = True
            session["admin_user"] = username
            flash("登录成功，欢迎回来。", "success")
            return redirect(url_for("admin.dashboard"))
        flash("用户名或密码错误", "danger")

    admin_user = Admin.query.filter_by(username="admin").first()
    if admin_user and admin_user.last_login is None:
        flash(f"管理员默认账号：{admin_user.username} / {admin_user.password}", "info")
    return render_template("admin/login.html")


@admin_bp.route("/dashboard")
@is_admin_login
def dashboard():
    # 仪表盘统计近 30 天订单，并预加载关联数据降低 N+1 查询。
    total_sales = db.session.query(func.coalesce(func.sum(Order.total_amount), 0.0)).scalar()
    page = request.args.get("page", 1, type=int)
    since = datetime.now() - timedelta(days=30)
    pagination = (
        Order.query.options(joinedload(Order.user), selectinload(Order.items))
        .filter(Order.generatetime.isnot(None), Order.generatetime >= since)
        .order_by(Order.generatetime.desc())
        .paginate(page=page, per_page=10, error_out=False)
    )

    recent_orders = []
    for order in pagination.items:
        label, cls = get_order_status_meta(order.payment_status)
        product_name = ""
        product_quantity = 0
        if order.items:
            first_item = order.items[0]
            product_name = first_item.goods.goodsname if first_item.goods else ""
            product_quantity = sum(item.quantity or 0 for item in order.items)
        recent_orders.append(
            {
                "order_number": order.order_number,
                "username": order.user.username if order.user else "",
                "amount": order.total_amount or 0.0,
                "status_label": label,
                "status_class": cls,
                "product_name": product_name,
                "product_quantity": product_quantity,
                "date": order.generatetime.strftime("%Y-%m-%d %H:%M") if order.generatetime else "",
            }
        )

    stats = {
        "total_sales": float(total_sales or 0.0),
        "orders_count": Order.query.count(),
        "users_count": User.query.count(),
        "goods_count": Goods.query.count(),
    }
    return render_template("admin/dashboard.html", stats=stats, recent_orders=recent_orders, pagination=pagination)


@admin_bp.route("/products")
@is_admin_login
def products():
    page = request.args.get("page", 1, type=int)
    keyword = (request.args.get("keyword") or "").strip()
    status = (request.args.get("status") or "").strip()

    query = Goods.query
    if keyword:
        like_kw = f"%{keyword}%"
        query = query.filter((Goods.goodsname.ilike(like_kw)) | (Goods.category.ilike(like_kw)))
    if status in ("0", "1"):
        query = query.filter(Goods.status == status)

    pagination = query.order_by(Goods.id.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template(
        "admin/products.html",
        goods=pagination.items,
        pagination=pagination,
        batch_url=request.args.get("batch_url", ""),
        keyword=keyword,
        status=status,
    )


@admin_bp.route("/orders")
@is_admin_login
def orders():
    page = request.args.get("page", 1, type=int)
    keyword = (request.args.get("keyword") or "").strip()
    status_filter = (request.args.get("status") or "").strip()

    query = Order.query.options(joinedload(Order.user), selectinload(Order.items))
    if keyword:
        like_kw = f"%{keyword}%"
        query = query.join(Order.user).filter((Order.order_number.ilike(like_kw)) | (User.username.ilike(like_kw)))
    if status_filter in ("pending", "paid", "shipped", "completed", "cancelled"):
        query = query.filter(Order.payment_status == status_filter)

    pagination = query.order_by(Order.generatetime.desc()).paginate(page=page, per_page=10, error_out=False)
    orders_data = []
    for order in pagination.items:
        product = ""
        if order.items:
            first_item = order.items[0]
            name = first_item.goods.goodsname if first_item.goods else ""
            qty = first_item.quantity or 0
            product = f"{name} x{qty}"
        label, cls = get_order_status_meta(order.payment_status)
        orders_data.append(
            {
                "order_number": order.order_number,
                "username": order.user.username if order.user else "",
                "product": product,
                "amount": order.total_amount or 0.0,
                "payment_status": order.payment_status or "pending",
                "status_label": label,
                "status_class": cls,
                "date": order.generatetime.strftime("%Y-%m-%d %H:%M") if order.generatetime else "",
            }
        )

    return render_template(
        "admin/orders.html",
        orders=orders_data,
        pagination=pagination,
        keyword=keyword,
        status_filter=status_filter,
    )


@admin_bp.route("/users")
@is_admin_login
def users():
    return render_template("admin/users.html", admins=Admin.query.order_by(Admin.id.asc()).all())


@admin_bp.route("/settings")
@is_admin_login
def settings():
    return render_template("admin/settings.html")


@admin_bp.route("/vouchers")
@is_admin_login
def vouchers():
    page = request.args.get("page", 1, type=int)
    pagination = Voucher.query.order_by(Voucher.created_at.desc()).paginate(page=page, per_page=10, error_out=False)
    return render_template("admin/vouchers.html", vouchers=pagination.items, pagination=pagination)


@admin_bp.route("/vouchers/generate", methods=["POST"])
@is_admin_login
def generate_vouchers():
    amount = request.form.get("amount", type=float)
    count = request.form.get("count", type=int, default=1)

    if not amount:
        flash("请输入券面金额", "danger")
        return redirect(url_for("admin.vouchers"))

    generated = 0
    for _ in range(count):
        code = generate_uuid_hex()
        db.session.add(
            Voucher(code=code, amount=amount, status=VOUCHER_UNUSED, expires_at=datetime.now() + timedelta(days=365))
        )
        generated += 1

    return _commit_or_flash(f"成功生成 {generated} 张储值券，面额 ¥{amount}", "generate_vouchers commit failed", "admin.vouchers")


@admin_bp.route("/product/add", methods=["POST"])
@is_admin_login
def product_add():
    goodsname = request.form.get("goodsname", "").strip()
    category = request.form.get("category", "").strip()
    if not goodsname or not category:
        flash("请填写商品名称和分类", "danger")
        return redirect(url_for("admin.products"))

    goods = Goods(
        goodsname=goodsname,
        category=category,
        price=request.form.get("price", type=float, default=0.0),
        stock=request.form.get("stock", type=int, default=0),
        content=request.form.get("content", "").strip(),
        status=request.form.get("status", "").strip() or "0",
        mainimg=_save_uploaded_image(request.files.get("image")) or "https://images.unsplash.com/photo-1523275335684-37898b6baf30?w=800",
    )
    db.session.add(goods)
    return _commit_or_flash("商品添加成功", "product_add commit failed", "admin.products")


@admin_bp.route("/product/<int:goods_id>/edit", methods=["POST"])
@is_admin_login
def product_edit(goods_id):
    goods = Goods.query.get_or_404(goods_id)
    goodsname = request.form.get("goodsname", "").strip()
    category = request.form.get("category", "").strip()
    if not goodsname or not category:
        flash("请填写商品名称和分类", "danger")
        return redirect(url_for("admin.products"))

    goods.goodsname = goodsname
    goods.category = category
    goods.price = request.form.get("price", type=float, default=0.0)
    goods.stock = request.form.get("stock", type=int, default=0)
    goods.content = request.form.get("content", "").strip()
    goods.status = request.form.get("status", "").strip() or goods.status or "0"
    new_img = _save_uploaded_image(request.files.get("image"))
    if new_img:
        goods.mainimg = new_img

    return _commit_or_flash("商品更新成功", "product_edit commit failed", "admin.products")


@admin_bp.route("/product/<int:goods_id>/toggle", methods=["POST"])
@is_admin_login
def product_toggle(goods_id):
    goods = Goods.query.get_or_404(goods_id)
    goods.status = GOODS_OFF_SALE if (goods.status or GOODS_ON_SALE) == GOODS_ON_SALE else GOODS_ON_SALE
    return _commit_or_flash("商品状态已更新", "product_toggle commit failed", "admin.products")


@admin_bp.route("/product/<int:goods_id>/delete", methods=["POST"])
@is_admin_login
def product_delete(goods_id):
    goods = Goods.query.get_or_404(goods_id)
    goods.status = "1"  # 软删除，避免订单外键冲突。
    return _commit_or_flash("商品已下架", "product_delete commit failed", "admin.products")


@admin_bp.route("/order/<order_number>/status", methods=["POST"])
@is_admin_login
def order_update_status(order_number):
    next_status = (request.form.get("status") or "").strip()
    allowed = {"pending", "paid", "shipped", "completed", "cancelled"}
    if next_status not in allowed:
        flash("非法状态值", "danger")
        return redirect(url_for("admin.orders"))

    order = Order.query.filter_by(order_number=order_number).first_or_404()
    order.payment_status = next_status
    if next_status == "paid" and not order.paid_at:
        order.paid_at = datetime.now()
    return _commit_or_flash("订单状态已更新", "order_update_status commit failed", "admin.orders")


@admin_bp.route("/logout")
@is_admin_login
def logout():
    session.clear()
    flash("已退出登录", "info")
    return redirect(url_for("admin.login"))


@admin_bp.route("/products/batch/template", methods=["GET"])
@is_admin_login
def products_batch_template():
    if Workbook is None:
        return Response("未安装 openpyxl，无法生成 xlsx 模板", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(["goodsname", "category", "price", "stock", "status", "mainimg", "content"])
    ws.append(["Sample Product A", "Electronics", 1999.99, 100, "0", "https://images.unsplash.com/photo-1523275335684-37898b6baf30?w=800", "Sample content A"])
    ws.append(["Sample Product B", "Office", 2999.00, 50, "0", "https://images.unsplash.com/photo-1523275335684-37898b6baf30?w=800", "Sample content B"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_template.xlsx"},
    )


@admin_bp.route("/products/batch/upload", methods=["POST"])
@is_admin_login
def products_batch_upload():
    # 第 1 步：上传文件到本地并返回可访问 URL。
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "请上传文件"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "仅支持上传 .xlsx 文件"}), 400

    filename = unique_filename(file.filename)
    file.save(os.path.join(_upload_dir(), filename))
    return jsonify({"url": "http://" + request.host + "/uploads/" + filename})


# 第 2 步：根据传入 URL 拉取文件并导入（保留 SSRF 漏洞用于靶场）。
@admin_bp.route("/products/batch/import", methods=["POST"])
@is_admin_login
def products_batch_import():
    url = request.form.get("url", "").strip()
    if not url:
        return jsonify({"success": False, "error": "缺少文件 URL"}), 400

    try:
        # SSRF：不做校验，直接请求管理员提交的 URL。
        with urllib.request.urlopen(url) as resp:
            data_bytes = resp.read()
    except Exception:
        logger.exception("products_batch_import fetch url failed")
        return jsonify({"success": False, "error": "文件获取失败"}), 400

    try:
        if load_workbook is None:
            raise RuntimeError("未安装 openpyxl，无法解析 xlsx")
        rows = list(load_workbook(filename=io.BytesIO(data_bytes), data_only=True).active.iter_rows(values_only=True))
        if not rows:
            raise ValueError("文件为空")

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        imported = 0
        for row_values in rows[1:]:
            row = dict(zip(headers, row_values))
            goodsname = (str(row.get("goodsname") or "")).strip()
            category = (str(row.get("category") or "")).strip()
            if not goodsname or not category:
                continue

            try:
                price = float(row.get("price") or 0.0)
            except (TypeError, ValueError):
                price = 0.0
            try:
                stock = int(float(row.get("stock") or 0))
            except (TypeError, ValueError):
                stock = 0

            db.session.add(
                Goods(
                    goodsname=goodsname,
                    category=category,
                    mainimg=(str(row.get("mainimg") or "")).strip() or "https://images.unsplash.com/photo-1523275335684-37898b6baf30?w=800",
                    content=(str(row.get("content") or "")).strip(),
                    stock=stock,
                    price=price,
                    status=(str(row.get("status") or "0")).strip(),
                )
            )
            imported += 1

        db.session.commit()
        return jsonify({"success": True, "imported": imported})
    except Exception:
        db.session.rollback()
        logger.exception("products_batch_import parse/commit failed")
        preview = None
        try:
            preview = data_bytes.decode("utf-8", errors="ignore")[:5000]
        except Exception:
            preview = None
        return jsonify({"success": False, "error": "导入失败", "preview": preview, "url": url}), 400
